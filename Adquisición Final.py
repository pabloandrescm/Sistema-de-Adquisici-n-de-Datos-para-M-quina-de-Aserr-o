# -*- coding: utf-8 -*-
"""
Monitor Industrial – adquisición multihilo, exportación, scroll,
tooltip, zoom/arrastre, visualización de energía y distancia.
Autor: Pablo · jun-2025
"""

from __future__ import annotations
import csv, os, struct, time, serial, shutil, re, threading, queue
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xlsxwriter
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from pymodbus.client import ModbusSerialClient
import math
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image as RLImage, Table as RLTable, TableStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from io import BytesIO
import matplotlib.pyplot as plt
import itertools

# --------------- Parámetros generales -----------------
PUERTO_PM,  BAUD_PM,  ESCLAVO_PM = "COM1", 38400, 1
PUERTO_ADAM, BAUD_ADAM           = "COM6", 9600
PUERTO_DIST, BAUD_DIST           = "COM2", 115200
MUESTREO_MS = 200   # 200 ms (5 Hz)
HIST_SEGS   = 60
TEMP_MAX    = 100
EXCEL_COLOR = "#1F497D"

REGS_PM = {
    "L1-L2 (V)": (3019, 3020), "L2-L3 (V)": (3021, 3022),
    "L3-L1 (V)": (3023, 3024), "L1-N (V)": (3027, 3028),
    "L2-N (V)": (3029, 3030), "L3-N (V)": (3031, 3032),
    "Potencia (kW)": (3059, 3060), "Potencia reactiva (kVAR)": (3067, 3068),
    "FP": (3191, 3192), "I1 (A)": (2999, 3000),
    "I2 (A)": (3001, 3002), "I3 (A)": (3003, 3004),
}

class MonitorIndustrial:
    _re_mm = re.compile(r"^\s*(\d+)\s*$")

    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("Monitor Industrial Aserrío")
        root.geometry("1150x1080")

        self.running = False
        self.after_id = None
        self.time_mode = 'rel'
        self.t0_dt = None
        self.track_active = self.track_anchor = False
        self.track_x = None
        self.t, self.I1, self.I2, self.I3, self.P, self.T, self.D = ([] for _ in range(7))
        self.pm_buf = {k: [] for k in REGS_PM}
        self.energy_kWh = 0.0

        self.base = self.csv_path = self.xl_path = ""
        self.csvf = self.csvw = self.xlwb = self.xlws = None
        self.data: dict[str, list[float]] = {}

        self.AUTO_SCALE_WINDOW_S = 15  # Ventana de autoescalado

        self._build_menu()
        self._build_gui()

        # Hilos y colas de datos
        self.q_pm    = queue.Queue()
        self.q_adam  = queue.Queue()
        self.q_dist  = queue.Queue()
        self.th_pm   = self.th_adam = self.th_dist = None

        # Estado de conexiones
        self.pm = None
        self.adam = None
        self.dist_ser = None
        
        self.tarifas = {
            "BT2": 170.88, "BT3": 170.88,
            "BT4.1": 170.64, "BT4.2": 170.64, "BT4.3": 170.64,
            "AT2": 166.44, "AT3": 166.44, "AT4.1/2/3": 166.44
        }

    # ========== MENÚ ==========
    def _build_menu(self):
        bar = tk.Menu(self.root)
        m_file = tk.Menu(bar, tearoff=0)
        m_file.add_command(label="Abrir CSV…",      command=self.import_csv)
        m_file.add_command(label="Exportar Excel…", command=self.export_to_excel)
        m_file.add_separator()
        m_file.add_command(label="Salir", command=self.root.quit)
        bar.add_cascade(label="Archivo", menu=m_file)
        m_track = tk.Menu(bar, tearoff=0)
        m_track.add_command(label="Activar",   command=lambda: self._set_track(True))
        m_track.add_command(label="Desactivar",command=lambda: self._set_track(False))
        bar.add_cascade(label="Barra de Seguimiento", menu=m_track)
        m_zoom = tk.Menu(bar, tearoff=0)
        m_zoom.add_command(label="Activar zoom + arrastre",   command=self.enable_zoom)
        m_zoom.add_command(label="Desactivar zoom + arrastre",command=self.disable_zoom)
        m_zoom.add_separator()
        m_zoom.add_command(label="Reestablecer vista", command=self.reset_view)
        bar.add_cascade(label="Visualización", menu=m_zoom)
        m_file.add_command(label="Generar Informe PDF…", command=self._generar_informe_pdf)

        self.root.config(menu=bar)

    # ========== INTERFAZ (gráficos, valores, scroll) ==========
    def _build_gui(self):
        style = ttk.Style(self.root)
        style.configure("Name.TLabel",  font=("Segoe UI", 11))
        style.configure("Value.TLabel", font=("Segoe UI", 14, "bold"))
        frm = ttk.LabelFrame(self.root, text="Valores instantáneos")
        frm.pack(fill="x", padx=6, pady=4)
        cols = 6
        self.lbl_val = {}
        for i, k in enumerate(REGS_PM):
            r, c = divmod(i, cols)
            ttk.Label(frm, text=k+":", style="Name.TLabel")\
                .grid(row=r, column=2*c, sticky="e", padx=(4,1), pady=2)
            lab = ttk.Label(frm, text="—", width=11, relief="groove",
                            style="Value.TLabel", anchor="center")
            lab.grid(row=r, column=2*c+1, sticky="we", padx=(0,8), pady=2)
            self.lbl_val[k] = lab
        r_fp, c_fp = divmod(list(REGS_PM).index("FP"), cols)
        tk.Button(frm, text="Iniciar Adquisición", command=self.start,
                  bg="#4CAF50", fg="white", activebackground="#45a049")\
            .grid(row=r_fp+1, column=2*c_fp-1, sticky="we", padx=(0,4))
        ttk.Label(frm, text="Temp (°C):", style="Name.TLabel")\
            .grid(row=r_fp+1, column=2*c_fp, sticky="e", padx=(4,1))
        self.lbl_temp = ttk.Label(frm, text="—", width=11, relief="groove",
                                  style="Value.TLabel", anchor="center")
        self.lbl_temp.grid(row=r_fp+1, column=2*c_fp+1, sticky="we")
        ttk.Label(frm, text="Dist (mm):", style="Name.TLabel")\
            .grid(row=r_fp+1, column=2*c_fp+2, sticky="e", padx=(4,1))
        self.lbl_dist = ttk.Label(frm, text="—", width=11, relief="groove",
                                  style="Value.TLabel", anchor="center")
        self.lbl_dist.grid(row=r_fp+1, column=2*c_fp+3, sticky="we")
        tk.Button(frm, text="Detener Adquisición", command=self.stop,
                  bg="#F44336", fg="white", activebackground="#d32f2f")\
            .grid(row=r_fp+1, column=2*c_fp+4, sticky="we", padx=(4,0))
        for c in range(cols*2+5):
            frm.columnconfigure(c, weight=1)

        self.fig = Figure(figsize=(10, 8.2), dpi=100)
        self.fig.subplots_adjust(top=0.96, bottom=0.13, left=0.07, right=0.98, hspace=0.33)
        self.ax_I = self.fig.add_subplot(4,1,1)
        self.ax_P = self.fig.add_subplot(4,1,2)
        self.ax_T = self.fig.add_subplot(4,1,3)
        self.ax_D = self.fig.add_subplot(4,1,4)
        self.lI1, = self.ax_I.plot([], [], label="I1", lw=2)
        self.lI2, = self.ax_I.plot([], [], label="I2", lw=2)
        self.lI3, = self.ax_I.plot([], [], label="I3", lw=2)
        self.lP , = self.ax_P.plot([], [], label="kW", lw=2)
        self.lT , = self.ax_T.plot([], [], label="Temp", color="red", lw=2)
        self.lD , = self.ax_D.plot([], [], label="Dist", color="purple", lw=2)
        for ax in (self.ax_I, self.ax_P, self.ax_T, self.ax_D):
            ax.grid(True); ax.margins(x=0); ax.set_xlim(0, HIST_SEGS)
        self.ax_I.set_ylim(0, 1); self.ax_P.set_ylim(0, 1)
        self.ax_T.set_ylim(0, TEMP_MAX); self.ax_D.set_ylim(0, 1)
        self.ax_I.set_title("Corrientes (A)");  self.ax_I.set_ylabel("A");  self.ax_I.legend()
        self.ax_P.set_title("Potencia (kW)");   self.ax_P.set_ylabel("kW"); self.ax_P.legend()
        self.ax_T.set_title("Temperatura (°C)");self.ax_T.set_ylabel("°C"); self.ax_T.legend()
        self.ax_D.set_title("Distancia (mm)");  self.ax_D.set_ylabel("mm"); self.ax_D.legend()
        self.ax_D.set_xlabel("Tiempo (s)")
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        # Scroll y energía abajo
        frame_bottom = ttk.Frame(self.root)
        frame_bottom.pack(fill="x", padx=20, pady=(2,8), side="bottom")
        self.lbl_E = ttk.Label(frame_bottom, text="Energía: 0.0000 kWh", style="Value.TLabel")
        self.lbl_E.pack(side="left", padx=(0,24))
        self.data_scroll = ttk.Scale(frame_bottom, orient="horizontal", from_=0, to=0,
                                     command=self._scroll_data, state="disabled")
        self.data_scroll.pack(side="left", fill="x", expand=True)

        # Seguimiento & tooltip
        self.v_I=self.ax_I.axvline(0,color="black",lw=2,zorder=10,visible=False)
        self.v_P=self.ax_P.axvline(0,color="black",lw=2,zorder=10,visible=False)
        self.v_T=self.ax_T.axvline(0,color="black",lw=2,zorder=10,visible=False)
        self.v_D=self.ax_D.axvline(0,color="black",lw=2,zorder=10,visible=False)
        self.hl_I=self.ax_I.scatter([],[],s=40,color="red",zorder=11,visible=False)
        self.hl_P=self.ax_P.scatter([],[],s=40,color="red",zorder=11,visible=False)
        self.hl_T=self.ax_T.scatter([],[],s=40,color="red",zorder=11,visible=False)
        self.hl_D=self.ax_D.scatter([],[],s=40,color="red",zorder=11,visible=False)
        self.txt_I=self.ax_I.text(0.02,0.95,"",transform=self.ax_I.transAxes,
                                  ha="left",va="top",fontsize=9,
                                  bbox=dict(fc="w",alpha=.8),visible=False)
        self.txt_P=self.ax_P.text(0.02,0.95,"",transform=self.ax_P.transAxes,
                                  ha="left",va="top",fontsize=9,
                                  bbox=dict(fc="w",alpha=.8),visible=False)
        self.txt_T=self.ax_T.text(0.02,0.95,"",transform=self.ax_T.transAxes,
                                  ha="left",va="top",fontsize=9,
                                  bbox=dict(fc="w",alpha=.8),visible=False)
        self.txt_D=self.ax_D.text(0.02,0.95,"",transform=self.ax_D.transAxes,
                                  ha="left",va="top",fontsize=9,
                                  bbox=dict(fc="w",alpha=.8),visible=False)
        self.annot=self.fig.text(0,0,"",va="bottom",ha="left",
                                 bbox=dict(fc="w",boxstyle="round",alpha=0.8),
                                 visible=False)
        self.canvas.mpl_connect("motion_notify_event", self._on_mouse_move)
        self.canvas.mpl_connect("button_press_event",  self._on_mouse_press)

    def _generar_informe_pdf(self):
        if not self.t:
            messagebox.showwarning("Informe PDF", "No hay datos para generar el informe.")
            return

        win = tk.Toplevel(self.root)
        win.title("Generar Informe PDF")
        win.grab_set()
        frame = ttk.Frame(win, padding=12)
        frame.pack()

        ttk.Label(frame, text="Seleccione tarifa para el cálculo de costos:").grid(row=0, column=0, columnspan=2, pady=(0, 10))

        tarifa_cb = ttk.Combobox(frame, width=15, state="readonly")
        tarifa_cb['values'] = list(self.tarifas.keys())
        tarifa_cb.current(0)
        tarifa_cb.grid(row=1, column=0, columnspan=2, pady=(0, 10))

        ttk.Button(frame, text="Editar tarifas", command=self.editar_tarifas_ventana).grid(row=2, column=0, columnspan=2, pady=6)

        def confirmar():
            tarifa = tarifa_cb.get()
            if tarifa not in self.tarifas:
                messagebox.showerror("Tarifa inválida", "Debe seleccionar una tarifa válida.")
                return
            carpeta = filedialog.askdirectory(title="Seleccionar carpeta de destino")
            if not carpeta:
                return  
            try:
                self._crear_pdf(carpeta, tarifa)
                messagebox.showinfo("Informe generado", "El informe se ha generado correctamente.")
                win.grab_release()
                win.destroy()
                if hasattr(self, "_ventana_edicion_tarifas") and self._ventana_edicion_tarifas.winfo_exists():
                    self._ventana_edicion_tarifas.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo generar el informe:\n{e}")

        ttk.Button(frame, text="Generar PDF", command=confirmar).grid(row=3, column=0, columnspan=2, pady=10)

    def editar_tarifas_ventana(self):
        if hasattr(self, "_ventana_edicion_tarifas") and self._ventana_edicion_tarifas.winfo_exists():
            self._ventana_edicion_tarifas.lift()
            return

        self._ventana_edicion_tarifas = tk.Toplevel(self.root)
        win = self._ventana_edicion_tarifas
        win.title("Editar tarifas")
        win.geometry("300x400")
        win.grab_set()

        ttk.Label(win, text="Editar tarifas (CLP/kWh):").pack(pady=10)

        entradas = {}

        frame_scroll = ttk.Frame(win)
        frame_scroll.pack(fill="both", expand=True, padx=10)

        canvas = tk.Canvas(frame_scroll)
        scrollbar = ttk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for i, (nombre, valor) in enumerate(self.tarifas.items()):
            ttk.Label(scrollable_frame, text=nombre + ":").grid(row=i, column=0, sticky="e", padx=4, pady=2)
            ent = ttk.Entry(scrollable_frame, width=10)
            ent.insert(0, f"{valor:.2f}")
            ent.grid(row=i, column=1, pady=2)
            entradas[nombre] = ent

        def guardar():
            for nombre, ent in entradas.items():
                try:
                    val = float(ent.get())
                    self.tarifas[nombre] = val
                except ValueError:
                    messagebox.showerror("Error", f"Valor inválido para {nombre}")
                    return
            if hasattr(self, "_tarifa_combobox") and self._tarifa_combobox:
                self._tarifa_combobox['values'] = list(self.tarifas.keys())
            messagebox.showinfo("Tarifas", "Tarifas actualizadas correctamente.")
            win.destroy()

        ttk.Button(win, text="Guardar cambios", command=guardar).pack(pady=10)

    # ========== MULTIHILO (Adquisición rápida y sincronizada) ==========
    def _connect_pm(self):
        self.pm=ModbusSerialClient(port=PUERTO_PM,baudrate=BAUD_PM,
                                   parity="E",stopbits=1,bytesize=8,timeout=0.1)
        return self.pm.connect()

    def _connect_adam(self):
        try:
            self.adam=serial.Serial(PUERTO_ADAM,BAUD_ADAM,bytesize=8,
                                    parity=serial.PARITY_NONE,stopbits=1,timeout=0.1)
            return True
        except serial.SerialException:
            return False

    def _connect_dist(self):
        try:
            self.dist_ser = serial.Serial(PUERTO_DIST, BAUD_DIST, timeout=0.1)
            self.dist_ser.reset_input_buffer()
            return True
        except serial.SerialException:
            return False

    def _read_pm(self):
        d = {}
        try:
            for k,p in REGS_PM.items():
                # Leer par de registros Modbus como float
                try:
                    r=self.pm.read_holding_registers(address=p[0],count=2,unit=ESCLAVO_PM)
                except TypeError:
                    r=self.pm.read_holding_registers(address=p[0],count=2,slave=ESCLAVO_PM)
                if getattr(r,"isError",lambda:True)():
                    return None
                d[k]=struct.unpack(">f",struct.pack(">HH",*r.registers))[0]
        except Exception:
            return None
        return d

    @staticmethod
    def _v_to_temp(v):    # 1-5 V → 0-100 °C
        return (v-1.0)/4.0*100.0

    def _read_adam(self):
        if not getattr(self,'adam',None) or not self.adam.is_open:
            return 0.0
        try:
            self.adam.reset_input_buffer(); self.adam.write(b"#02\r"); time.sleep(0.1)
            resp=self.adam.read(self.adam.in_waiting).decode(errors="ignore").strip()
            v=float(resp.split('+')[-1]) if resp else 1.0
            return self._v_to_temp(v)
        except Exception:
            return 0.0

    def _read_dist(self):
        if not getattr(self, "dist_ser", None) or not self.dist_ser.is_open:
            return 0.0
        last = getattr(self, "_last_dist_mm", 0.0)
        try:
            while self.dist_ser.in_waiting > 0:
                line = self.dist_ser.readline().decode("utf-8", errors="ignore").strip()
                m = self._re_mm.match(line)
                if m:
                    last = float(m.group(1))
            self._last_dist_mm = last
        except Exception:
            pass
        return last

    # ========== Hilos ==========
    def _thread_pm(self):
        while self.running:
            dpm = self._read_pm()
            if dpm is not None:
                self.q_pm.put((time.time(), dpm))
            time.sleep(MUESTREO_MS / 1000.0)

    def _thread_adam(self):
        while self.running:
            t = self._read_adam()
            self.q_adam.put((time.time(), t))
            time.sleep(MUESTREO_MS / 1000.0)

    def _thread_dist(self):
        while self.running:
            dist = self._read_dist()
            now = time.time()
            # Actualiza la GUI en tiempo real solo el valor instantáneo
            self.root.after(0, self.lbl_dist.config, {"text": f"{dist:.0f}"})
            # Encola para sincronización y registro
            self.q_dist.put((now, dist))
            time.sleep(MUESTREO_MS / 1000.0)

    # ========== Control de adquisición (start/stop/loop) ==========
    def start(self):
        if self.running: return
        if not self._connect_pm():
            messagebox.showerror("PM2120","No se pudo conectar"); return
        if not self._connect_adam():
            messagebox.showerror("ADAM-4017","No se pudo abrir puerto"); self.pm.close(); return
        if not self._connect_dist():
            messagebox.showerror("Distancia","No se pudo abrir puerto COM2"); 
            self.pm.close(); self.adam.close(); return

        # --- Fija t0_dt para coherencia temporal en outputs ---
        self.t0_dt = datetime.now()

        self._open_files()
        self.running=True; self.time_mode='rel'
        self.start_ts=self.last_ts=time.time(); self.energy_kWh=0.0
        for lst in (self.t,self.I1,self.I2,self.I3,self.P,self.T,self.D): lst.clear()
        for lst in self.pm_buf.values(): lst.clear()
        self.track_x=None; self.track_anchor=False
        self.data_scroll.configure(state="disabled")

        # Iniciar hilos
        self.q_pm.queue.clear(); self.q_adam.queue.clear(); self.q_dist.queue.clear()
        self.th_pm = threading.Thread(target=self._thread_pm, daemon=True)
        self.th_adam = threading.Thread(target=self._thread_adam, daemon=True)
        self.th_dist = threading.Thread(target=self._thread_dist, daemon=True)
        self.th_pm.start(); self.th_adam.start(); self.th_dist.start()

        self._main_loop()

    def stop(self):
        if not self.running: return
        self.running=False
        if self.after_id: self.root.after_cancel(self.after_id)
        try:
            if self.pm and self.pm.connected: self.pm.close()
        except Exception: pass
        try:
            if getattr(self,'adam',None) and self.adam.is_open: self.adam.close()
        except Exception: pass
        try:
            if getattr(self,'dist_ser',None) and self.dist_ser.is_open: self.dist_ser.close()
        except Exception: pass

        if self.csvf: self.csvf.close()

        # --- Construir Excel completo desde el CSV recién creado ---
        try:
            self._finalize_excel_from_csv()   # llena hoja "Datos" leyendo el CSV
            self._add_charts()                # crea hoja "Gráficas"
            self.xlwb.save(self.xl_path)      # guarda el .xlsx completo
        except Exception as e:
            messagebox.showerror("Excel", f"Error al finalizar Excel:\n{e}")

        dest=filedialog.askdirectory(title="Carpeta destino")
        if dest:
            try:
                shutil.move(self.csv_path,os.path.join(dest,os.path.basename(self.csv_path)))
                shutil.move(self.xl_path, os.path.join(dest,os.path.basename(self.xl_path)))
                msg=f"Archivos guardados en:\n{dest}"
            except Exception as e:
                msg=f"No se pudieron mover los archivos:\n{e}"
        else:
            msg="No se eligió carpeta.\nLos archivos quedaron en la carpeta actual."

        self._enable_scroll()
        messagebox.showinfo("Fin de adquisición",
                            f"Energía total: {self.energy_kWh:.4f} kWh.\n{msg}")

    # ========== LOOP PRINCIPAL ==========
    def _main_loop(self):
        if not self.running:
            return

        # Función auxiliar: tomar solo el dato más reciente
        def _get_latest(q):
            item = None
            while not q.empty():
                item = q.get()
            return item

        dpm_data = _get_latest(self.q_pm)
        adam_data = _get_latest(self.q_adam)
        dist_data = _get_latest(self.q_dist)

        # Si falta algún dato, esperar siguiente ciclo
        if not (dpm_data and adam_data and dist_data):
            self.after_id = self.root.after(MUESTREO_MS, self._main_loop)
            return

        ts_pm, dpm = dpm_data
        ts_adam, temp = adam_data
        ts_dist, dist = dist_data

        now = time.time()

        # Actualizar etiquetas en GUI
        for k, v in dpm.items():
            self.lbl_val[k].config(text=f"{v:.3f}")
        self.lbl_temp.config(text=f"{temp:.2f}")
        self.lbl_dist.config(text=f"{dist:.0f}")

        # Agregar datos nuevos a listas
        t_rel = now - self.start_ts
        self.t.append(t_rel)

        for k in REGS_PM:
            self.pm_buf[k].append(dpm[k])
        self.I1.append(dpm["I1 (A)"])
        self.I2.append(dpm["I2 (A)"])
        self.I3.append(dpm["I3 (A)"])
        self.P.append(dpm["Potencia (kW)"])
        self.T.append(temp)
        self.D.append(dist)

        # Actualizar curvas
        self.lI1.set_data(self.t, self.I1)
        self.lI2.set_data(self.t, self.I2)
        self.lI3.set_data(self.t, self.I3)
        self.lP.set_data(self.t, self.P)
        self.lT.set_data(self.t, self.T)
        self.lD.set_data(self.t, self.D)

        for ax in (self.ax_I, self.ax_P, self.ax_T, self.ax_D):
            ax.set_xlim(max(0, t_rel - HIST_SEGS), t_rel)

        # Actualizar gráficos cada 3 ciclos
        self._plot_counter = getattr(self, '_plot_counter', 0) + 1
        if self._plot_counter % 3 == 0:
            self._autoscale_axes()
            self.canvas.draw_idle()

        # Calcular energía
        self.energy_kWh += dpm["Potencia (kW)"] * (now - self.last_ts) / 3600
        self.last_ts = now
        self.lbl_E.config(text=f"Energía: {self.energy_kWh:.4f} kWh")

        # Registrar en CSV
        ts_iso = datetime.now().isoformat(' ', timespec='seconds')
        fila = [ts_iso] + [f"{dpm[k]:.3f}" for k in REGS_PM] + [f"{temp:.2f}", f"{dist:.2f}"]
        self.csvw.writerow(fila)

        # Agendar próximo ciclo
        self.after_id = self.root.after(MUESTREO_MS, self._main_loop)

    # ========== AUTOSCALE DINÁMICO ==========
    def _autoscale_axes(self):
        if not self.t:
            return
        # Ejes: I1/I2/I3, P, T, D. Escala eje Y SOLO para los datos visibles
        for ax, vals_list in [(self.ax_I, [self.I1, self.I2, self.I3]),
                              (self.ax_P, [self.P]),
                              (self.ax_T, [self.T]),
                              (self.ax_D, [self.D])]:
            xmin, xmax = ax.get_xlim()
            if self.t:
                idx0 = next((i for i, tx in enumerate(self.t) if tx >= xmin), len(self.t)-1)
                idx1 = next((i for i, tx in enumerate(self.t) if tx > xmax), len(self.t))
            else:
                idx0, idx1 = 0, 0
            window_vals = []
            for vals in vals_list:
                window_vals.extend(vals[idx0:idx1])
            if ax in (self.ax_I, self.ax_P):
                ymax = max(window_vals) if window_vals else 1.0
                ax.set_ylim(0, max(ymax * 1.15, 1.0))
            elif ax is self.ax_T:
                ymax = max(window_vals) if window_vals else TEMP_MAX
                ax.set_ylim(0, max(ymax * 1.05, TEMP_MAX))
            else:
                ymax = max(window_vals) if window_vals else 1.0
                ax.set_ylim(0, max(ymax * 1.15, 1.0))

    def reset_view(self):
        if not self.t:
            return
        t0 = self.t[0]
        t1 = self.t[-1]
        for ax in (self.ax_I, self.ax_P, self.ax_T, self.ax_D):
            ax.set_xlim(t0, t1)
        self._autoscale_axes()
        self.canvas.draw_idle()

    # ========== SCROLL ==========
    def _scroll_data(self, val):
        if not self.t: return
        span = HIST_SEGS if self.time_mode == 'rel' else (self.t[-1] - self.t[0])
        t0 = float(val)
        t1 = t0 + span
        for ax in (self.ax_I, self.ax_P, self.ax_T, self.ax_D):
            ax.set_xlim(t0, t1)
        self._autoscale_axes()
        self.canvas.draw_idle()

    def _enable_scroll(self):
        if not self.t: return
        span = HIST_SEGS if self.time_mode == 'rel' else (self.t[-1] - self.t[0])
        mx = max(0, self.t[-1] - span)
        self.data_scroll.configure(to=mx, state="normal"); self.data_scroll.set(mx)

    # ========== ARCHIVOS Y EXPORTACION ==========
    def _open_files(self):
        self.base = datetime.now().strftime("adq_%Y%m%d_%H%M%S")
        hdr = ["timestamp_iso"] + list(REGS_PM.keys()) + ["Temp (°C)", "Dist (mm)"]
        self.csv_path = self.base + ".csv"
        self.xl_path = self.base + ".xlsx"
        self.csvf = open(self.csv_path, "w", newline="", encoding="utf-8")
        self.csvw = csv.writer(self.csvf)
        self.csvw.writerow(hdr)
        self.xlwb = Workbook()
        self.xlws = self.xlwb.active
        self.xlws.title = "Datos"
        self.xlws.append(hdr)

    # --- llenar Excel desde el CSV recién generado ---
    def _finalize_excel_from_csv(self):
        """
        Llena la hoja 'Datos' del workbook (openpyxl) leyendo el CSV recién generado.
        No escribe Excel en tiempo real: solo al finalizar.
        """
        if not os.path.exists(self.csv_path):
            raise FileNotFoundError(f"CSV no encontrado: {self.csv_path}")

        ws = self.xlws

        # Deja sólo la cabecera actual
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        with open(self.csv_path, newline="", encoding="utf-8") as fp:
            rdr = csv.reader(fp)
            hdr = next(rdr, None)
            if not hdr:
                return

            # Asegura cabecera correcta
            ws.delete_rows(1, 1)
            ws.append(hdr)

            for row in rdr:
                if not row:
                    continue
                out = [row[0]]  # timestamp como texto
                for cell in row[1:]:
                    try:
                        out.append(float(cell.replace(",", ".").strip()))
                    except Exception:
                        out.append(0.0)
                ws.append(out)

    # --- AJUSTE CHARTS: usa Reference con argumentos nombrados y categorías correctas ---
    def _add_charts(self):
        ws_d = self.xlws
        ws_g = self.xlwb.create_sheet("Gráficas")
        max_r = ws_d.max_row
        max_c = ws_d.max_column
        if max_r < 2 or max_c < 2:
            return  # no hay datos

        # Categorías: columna 1 (timestamp_iso), filas desde 2 hasta max_r
        ts_ref = Reference(ws_d, min_col=1, min_row=2, max_col=1, max_row=max_r)

        row_anchor = 1  # fila donde vamos poniendo gráficos en "Gráficas"
        for col in range(2, max_c + 1):
            hdr = ws_d.cell(row=1, column=col).value or f"Col {col}"
            data_ref = Reference(ws_d, min_col=col, min_row=1, max_col=col, max_row=max_r)

            ch = LineChart()
            ch.title = hdr
            ch.y_axis.title = hdr
            ch.x_axis.title = "Tiempo"
            ch.add_data(data_ref, titles_from_data=True)
            ch.set_categories(ts_ref)
            ch.height = 15  # ~ 7 cm
            ch.width  = 32  # ~ 16 cm

            ws_g.add_chart(ch, f"A{row_anchor}")
            row_anchor += 16  

    # ========== IMPORTAR Y EXPORTAR ==========
    def import_csv(self):
        if self.running:
            messagebox.showinfo("Adquisición activa", "Detén la adquisición antes de importar.")
            return
        fname = filedialog.askopenfilename(title="Selecciona CSV",
                                           filetypes=[("Archivos CSV", "*.csv")])
        if not fname:
            return
        try:
            with open(fname, newline="", encoding="utf-8") as fp:
                rdr = csv.reader(fp)
                hdr = next(rdr)
                if "timestamp_iso" not in hdr:
                    raise ValueError("Falta la columna «timestamp_iso» en el encabezado.")
                idx_ts = hdr.index("timestamp_iso")
                self.data = {h: [] for h in hdr if h != "timestamp_iso"}
                self.t.clear(); self.I1.clear(); self.I2.clear()
                self.I3.clear(); self.P.clear(); self.T.clear(); self.D.clear()
                self.t0_dt = None
                dup = {}
                for n_row, row in enumerate(rdr, start=2):
                    if not row or all(c.strip() == "" for c in row):
                        continue
                    if len(row) < len(hdr):
                        row += [""] * (len(hdr) - len(row))
                    ts = row[idx_ts].strip()
                    if not ts:
                        continue
                    try:
                        dt = datetime.fromisoformat(ts)
                    except Exception:
                        continue
                    self.t0_dt = self.t0_dt or dt
                    sec = (dt - self.t0_dt).total_seconds()
                    sec += dup.get(int(sec), 0) * 1e-4
                    dup[int(sec)] = dup.get(int(sec), 0) + 1
                    self.t.append(sec)
                    for i, h in enumerate(hdr):
                        if i == idx_ts:
                            continue
                        txt = row[i].replace(",", ".").strip()
                        try:
                            val = float(txt) if txt else 0.0
                        except ValueError:
                            val = 0.0
                        self.data[h].append(val)

                rows = len(self.t)
                col = lambda name, d=0.0: self.data.get(name, [d]*rows)
                self.I1[:] = col("I1 (A)")
                self.I2[:] = col("I2 (A)")
                self.I3[:] = col("I3 (A)")
                self.P[:]  = col("Potencia (kW)")
                temp_cols  = [n for n in self.data if "temp" in n.lower()]
                self.T[:]  = self.data[temp_cols[0]] if temp_cols else [0.0]*rows
                dist_cols  = [n for n in self.data if "dist" in n.lower()]
                self.D[:]  = self.data[dist_cols[0]] if dist_cols else [0.0]*rows

                if not self.t:
                    messagebox.showwarning("Importar CSV", "No se encontraron datos válidos.")
                    return

                self.time_mode = 'dt'
                self.track_x = self.track_anchor = None
                self._refresh_plot()
                self._enable_scroll()

                # Calcular energía robustamente
                try:
                    potencias = self.data.get("Potencia (kW)", [])
                    energia_kWh = 0.0
                    for i in range(1, len(self.t)):
                        p = potencias[i]
                        if p <= 0:
                            continue  
                        dt_horas = max((self.t[i] - self.t[i - 1]) / 3600.0, 0)
                        energia_kWh += p * dt_horas
                    self.energy_kWh = energia_kWh
                    self.lbl_E.config(text=f"Energía: {self.energy_kWh:.4f} kWh")
                except Exception as e:
                    self.lbl_E.config(text="Energía: 0.0000 kWh")
                    print("Error al calcular energía:", e)
                    messagebox.showwarning("Cálculo de energía", f"No se pudo calcular energía:\n{e}")

        except Exception as e:
            messagebox.showerror("Importar CSV", f"No se pudo importar:\n{e}")

    def export_to_excel(self):
        if not self.t:
            messagebox.showinfo("Exportar", "No hay datos para exportar.")
            return
        fname = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Libro Excel", "*.xlsx")]
        )
        if not fname:
            return
        try:
            wb = xlsxwriter.Workbook(fname)
            ws = wb.add_worksheet("Datos")
            hdr = ["Hora"] + list(self.data.keys())
            ws.write_row(0, 0, hdr)

            # --- LIMPIA NaN e Inf en los datos ---
            clean_data = {}
            for key, vals in self.data.items():
                clean_vals = []
                for v in vals:
                    try:
                        v_float = float(v)
                    except Exception:
                        v_float = 0.0
                    if math.isnan(v_float) or math.isinf(v_float):
                        v_float = 0.0
                    clean_vals.append(v_float)
                clean_data[key] = clean_vals

            for i, tseg in enumerate(self.t, start=1):
                hora = (self.t0_dt + timedelta(seconds=tseg)).strftime("%H:%M:%S.%f")[:-3] \
                    if self.time_mode == 'dt' else f"{tseg:.3f}"
                ws.write_string(i, 0, hora)
                for c, h in enumerate(clean_data, start=1):
                    ws.write_number(i, c, clean_data[h][i - 1])

            cats = ['Datos', 1, 0, len(self.t), 0]
            ws_g = wb.add_worksheet("Graficas")
            for k, (n, col_vals) in enumerate(clean_data.items()):
                ch = wb.add_chart({'type': 'line'})
                ch.add_series({
                    'categories': cats,
                    'values': ['Datos', 1, k + 1, len(self.t), k + 1],
                    'name': n,
                    'line': {'color': EXCEL_COLOR}
                })
                ch.set_title({'name': n})
                ch.set_x_axis({'name': 'Hora'})
                ch.set_y_axis({'name': n, 'min': 0, 'max': max(col_vals) * 1.15 if col_vals else 1})
                ch.set_size({'width': 1600, 'height': 600})
                ws_g.insert_chart(k * 30, 0, ch)
            wb.close()
            messagebox.showinfo("Exportar", "Excel exportado correctamente.")
        except Exception as e:
            messagebox.showerror("Exportar", f"No se pudo crear Excel:\n{e}")

    # ========== PLOT REFRESH ==========
    def _refresh_plot(self):
        self.lI1.set_data(self.t, self.I1)
        self.lI2.set_data(self.t, self.I2)
        self.lI3.set_data(self.t, self.I3)
        self.lP .set_data(self.t, self.P)
        self.lT .set_data(self.t, self.T)
        self.lD .set_data(self.t, self.D)
        if self.t:
            xmin, xmax = self.t[0], self.t[-1]
            for ax in (self.ax_I, self.ax_P, self.ax_T, self.ax_D):
                ax.set_xlim(xmin, xmax)
        self._autoscale_axes()
        self.canvas.draw_idle()

    # ========== ZOOM/TOOLTIP ==========
    def enable_zoom(self):
        if getattr(self, "zoom_enabled", False): return
        self._zoom_cid = self.canvas.mpl_connect('scroll_event', self._on_zoom)
        self._pan_press_cid  = self.canvas.mpl_connect('button_press_event',  self._on_pan_press)
        self._pan_motion_cid = self.canvas.mpl_connect('motion_notify_event', self._on_pan_motion)
        self._pan_release_cid= self.canvas.mpl_connect('button_release_event',self._on_pan_release)
        self.zoom_enabled=True

    def disable_zoom(self):
        if not getattr(self, "zoom_enabled", False): return
        self.canvas.mpl_disconnect(self._zoom_cid)
        self.canvas.mpl_disconnect(self._pan_press_cid)
        self.canvas.mpl_disconnect(self._pan_motion_cid)
        self.canvas.mpl_disconnect(self._pan_release_cid)
        self.zoom_enabled=False; self.pan_active=False

    def _on_zoom(self, event):
        if event.inaxes not in (self.ax_I, self.ax_P, self.ax_T, self.ax_D): return
        ax = event.inaxes
        x0, x1 = ax.get_xlim(); y0, y1 = ax.get_ylim()
        xc = event.xdata if event.xdata is not None else (x0 + x1) / 2
        yc = event.ydata if event.ydata is not None else (y0 + y1) / 2
        s = 1.2 if event.button == 'up' else 0.8
        ax.set_xlim(xc - (xc - x0) / s, xc + (x1 - xc) / s)
        ax.set_ylim(yc - (yc - y0) / s, yc + (y1 - yc) / s)
        if ax in (self.ax_I, self.ax_P) and ax.get_ylim()[0] < 0:
            yb, yt = ax.get_ylim(); ax.set_ylim(0, max(0.01, yt))
        self._autoscale_axes()
        self.canvas.draw_idle()

    def _on_pan_press(self, event):
        if not getattr(self, "zoom_enabled", False) or event.button != 1: return
        if event.inaxes not in (self.ax_I, self.ax_P, self.ax_T, self.ax_D): return
        self.pan_active = True
        self.pan_ax = event.inaxes
        self.pan_last = (event.xdata, event.ydata)

    def _on_pan_motion(self, event):
        if not (getattr(self, "pan_active", False) and event.inaxes == self.pan_ax and event.xdata is not None):
            return
        dx = self.pan_last[0] - event.xdata
        dy = self.pan_last[1] - event.ydata
        x0, x1 = self.pan_ax.get_xlim(); y0, y1 = self.pan_ax.get_ylim()
        self.pan_ax.set_xlim(x0 + dx, x1 + dx)
        self.pan_ax.set_ylim(y0 + dy, y1 + dy)
        if self.pan_ax in (self.ax_I, self.ax_P) and self.pan_ax.get_ylim()[0] < 0:
            yb, yt = self.pan_ax.get_ylim(); self.pan_ax.set_ylim(0, max(0.01, yt))
        self.pan_last = (event.xdata, event.ydata)
        self._autoscale_axes()
        self.canvas.draw_idle()

    def _on_pan_release(self, event):
        if event.button == 1:
            self.pan_active = False

    def _on_mouse_press(self, event):
        if event.button==3 and self.track_active:
            self.track_anchor = not self.track_anchor

    def _on_mouse_move(self, event):
        if (self.track_active and not self.track_anchor and
            event.inaxes in (self.ax_I,self.ax_P,self.ax_T,self.ax_D) and event.xdata is not None):
            self._move_track(event.xdata)
        self._show_tooltip(event)

    def _show_tooltip(self, event):
        if self.track_anchor:
            self.annot.set_visible(False)
            self.canvas.draw_idle()
            return
        ax_map = {
            self.ax_I: (self.hl_I, [self.lI1, self.lI2, self.lI3]),
            self.ax_P: (self.hl_P, [self.lP]),
            self.ax_T: (self.hl_T, [self.lT]),
            self.ax_D: (self.hl_D, [self.lD])
        }
        if event.inaxes not in ax_map:
            self.annot.set_visible(False)
            for hl in (self.hl_I, self.hl_P, self.hl_T, self.hl_D):
                hl.set_visible(False)
            self.canvas.draw_idle()
            return
        hl, lines = ax_map[event.inaxes]
        xq = event.xdata
        best = None
        dmin = float("inf")
        for ln in lines:
            xs = ln.get_xdata()
            if hasattr(xs, 'size'):
                if xs.size == 0: continue
            elif not xs: continue
            idx = min(range(len(xs)), key=lambda i: abs(xs[i] - xq))
            d = abs(xs[idx] - xq)
            if d < dmin:
                dmin, best = d, (ln, idx)
        if not best:
            return
        ln, idx = best
        xv = ln.get_xdata()[idx]
        if event.inaxes is self.ax_I:
            txt = (f"I1={self.I1[idx]:.2f} A\n"
                   f"I2={self.I2[idx]:.2f} A\n"
                   f"I3={self.I3[idx]:.2f} A")
        elif event.inaxes is self.ax_P:
            txt = f"{self.P[idx]:.2f} kW"
        elif event.inaxes is self.ax_T:
            txt = f"{self.T[idx]:.2f} °C"
        elif event.inaxes is self.ax_D:
            txt = f"{self.D[idx]:.0f} mm"
        else:
            txt = f"{ln.get_label()}\nY={ln.get_ydata()[idx]:.4f}"
        tstr = (f"{xv:.3f}s" if self.time_mode == 'rel'
                else (self.t0_dt + timedelta(seconds=xv)).strftime("%H:%M:%S.%f")[:-3])
        self.annot.set_text(f"{tstr}\n{txt}")
        xf, yf = self.fig.transFigure.inverted().transform(
            event.inaxes.transData.transform((xv, ln.get_ydata()[idx])))
        self.annot.set_position((xf, yf))
        self.annot.set_visible(True)
        hl.set_offsets([[xv, ln.get_ydata()[idx]]])
        hl.set_visible(True)
        self.canvas.draw_idle()

    def _set_track(self,val:bool):
        self.track_active = val; self.track_anchor=False
        for v in (self.v_I,self.v_P,self.v_T,self.v_D): v.set_visible(val)
        for t in (self.txt_I,self.txt_P,self.txt_T,self.txt_D): t.set_visible(val)
        if val:
            if self.track_x is None:
                xmin,xmax=self.ax_I.get_xlim(); self.track_x=0.5*(xmin+xmax) if self.t else 0.0
            self._move_track(self.track_x)
        self.canvas.draw_idle()

    def _move_track(self,x:float):
        self.track_x=x
        for v in (self.v_I,self.v_P,self.v_T,self.v_D): v.set_xdata([x,x])
        if self.t:
            idx=min(range(len(self.t)),key=lambda i:abs(self.t[i]-x))
            time_str=(f"{self.t[idx]:.3f}s" if self.time_mode=='rel'
                      else (self.t0_dt+timedelta(seconds=self.t[idx])).strftime("%H:%M:%S.%f")[:-3])
            self.hl_I.set_offsets([[self.t[idx],self.I1[idx]]])
            self.hl_P.set_offsets([[self.t[idx],self.P[idx]]])
            self.hl_T.set_offsets([[self.t[idx],self.T[idx]]])
            self.hl_D.set_offsets([[self.t[idx],self.D[idx]]])
            self.hl_I.set_visible(True); self.hl_P.set_visible(True)
            self.hl_T.set_visible(True); self.hl_D.set_visible(True)
            self.txt_I.set_text(f"{time_str}\nI1={self.I1[idx]:.2f}\nI2={self.I2[idx]:.2f}\nI3={self.I3[idx]:.2f}")
            self.txt_P.set_text(f"{time_str}\nP={self.P[idx]:.2f} kW")
            self.txt_T.set_text(f"{time_str}\nT={self.T[idx]:.2f} °C")
            self.txt_D.set_text(f"{time_str}\nD={self.D[idx]:.0f} mm")
        self.canvas.draw_idle()

    def _crear_pdf(self, carpeta, tarifa):
        try:
            nombre = datetime.now().strftime("informe_aserrio_%Y%m%d_%H%M%S") + ".pdf"
            ruta = os.path.join(carpeta, nombre)
            doc = SimpleDocTemplate(ruta, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # ─── Portada ─────────────────────────────────────────────
            story.append(Spacer(1, 5*cm))
            story.append(Paragraph("<b>Informe de adquisición de datos de aserrío</b>", styles["Title"]))
            story.append(Spacer(1, 1*cm))

            if hasattr(self, "t0_dt") and self.t0_dt and self.t:
                dt_inicio = self.t0_dt
                dt_fin = self.t0_dt + timedelta(seconds=self.t[-1])
                duracion = dt_fin - dt_inicio
                dur_texto = str(duracion).split(".")[0]  

                tarifa_valor = self.tarifas.get(tarifa, 0)
                energia = self.energy_kWh
                costo_total = energia * tarifa_valor

                texto_info = f"""
                <para align='center'>
                Inicio adquisición: {dt_inicio.strftime('%d-%m-%Y %H:%M:%S')}<br/>
                Fin adquisición: {dt_fin.strftime('%d-%m-%Y %H:%M:%S')}<br/>
                Duración total: {dur_texto}<br/><br/>
                Tarifa seleccionada: <b>{tarifa}</b> → ${tarifa_valor:.2f} CLP/kWh<br/>
                Energía total: {energia:.4f} kWh<br/>
                <b>Costo total estimado:</b> ${costo_total:,.0f} CLP
                </para>
                """
            else:
                texto_info = "<para align='center'>Fecha de adquisición no disponible</para>"

            story.append(Paragraph(texto_info, styles["Normal"]))
            story.append(PageBreak())

            # ─── Gráficos ────────────────────────────────────────────
            columnas = list(self.data.keys())
            color_cycle = itertools.cycle(plt.rcParams['axes.prop_cycle'].by_key()['color'])

            for i in range(0, len(columnas), 3):
                fig, axs = plt.subplots(nrows=3, ncols=1, figsize=(8, 10))
                for j in range(3):
                    if i + j >= len(columnas): break
                    col = columnas[i + j]
                    color = next(color_cycle)
                    axs[j].plot(self.t, self.data[col], label=col, color=color)
                    axs[j].set_title(col, fontsize=10, fontweight='bold')
                    axs[j].set_xlabel("Tiempo (s)" if self.time_mode == "rel" else "Tiempo(s)")
                    axs[j].set_ylabel(col)
                    axs[j].legend()
                    axs[j].grid(True)
                fig.tight_layout()
                buffer = BytesIO()
                fig.savefig(buffer, format="PNG", dpi=150)
                plt.close(fig)
                buffer.seek(0)
                story.append(RLImage(buffer, width=16*cm, height=20*cm))
                story.append(PageBreak())

            # ─── Tabla de datos ──────────────────────────────────────
            hdr = ["Tiempo"] + columnas
            tabla_data = [hdr]
            for i in range(len(self.t)):
                fila = [f"{self.t[i]:.3f}" if self.time_mode == 'rel'
                        else (self.t0_dt + timedelta(seconds=self.t[i])).strftime("%H:%M:%S")]
                fila += [f"{self.data[c][i]:.2f}" for c in columnas]
                tabla_data.append(fila)

            tabla = RLTable(tabla_data, repeatRows=1, colWidths=[A4[0] / len(tabla_data[0])] * len(tabla_data[0]))
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
            ]))
            story.append(Paragraph("Tabla completa de datos:", styles["Heading2"]))
            story.append(tabla)
            story.append(PageBreak())

            doc.build(story)

        except Exception as e:
            raise RuntimeError(f"Error al crear el informe PDF: {e}")

# ========== EJECUCIÓN ==========
if __name__ == "__main__":
    os.environ.setdefault("MPLBACKEND", "TkAgg")
    root = tk.Tk(); MonitorIndustrial(root); root.mainloop()
